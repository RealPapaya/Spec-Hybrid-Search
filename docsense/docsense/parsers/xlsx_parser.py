"""
XLSX parser — openpyxl.

Reads every worksheet in the workbook.  Each sheet becomes a single
``table``-typed :class:`TextBlock` with the sheet name used as section
title.  Rows are rendered as pipe-separated strings so they remain
human-readable in search excerpts.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from docsense.parsers.base import ParsedDocument, ParserRegistry, TextBlock

logger = logging.getLogger(__name__)


class XlsxParser:
    """Parse ``.xlsx`` files into :class:`TextBlock` objects (one per sheet)."""

    def parse(self, filepath: Path) -> ParsedDocument:
        """
        Extract text from all worksheets in an Excel workbook.

        Parameters
        ----------
        filepath:
            Path to the ``.xlsx`` file.

        Returns
        -------
        ParsedDocument
            One block per non-empty worksheet.
        """
        blocks: list[TextBlock] = []

        try:
            wb = load_workbook(str(filepath), read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_text: list[str] = []

                for row in ws.iter_rows(values_only=True):
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    row_str = " | ".join(c for c in cells if c.strip())
                    if row_str.strip():
                        rows_text.append(row_str)

                if rows_text:
                    blocks.append(TextBlock(
                        text="\n".join(rows_text),
                        section_title=f"Sheet: {sheet_name}",
                        block_type="table",
                    ))

            wb.close()

        except Exception as exc:
            logger.error("XLSX parse error for %s: %s", filepath, exc)
            return ParsedDocument(
                filename=filepath.name,
                file_type=".xlsx",
                error=str(exc),
            )

        return ParsedDocument(
            filename=filepath.name,
            file_type=".xlsx",
            blocks=blocks,
        )


# Register with the global parser registry
ParserRegistry.register(".xlsx", XlsxParser())
